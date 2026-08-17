#!/usr/bin/env python3
"""Set 看板-单城 C2/C3/E3 via openpyxl, then write a one-city WPS export JSON.

WPS COM must not write these cells (32-bit KET Value2 is Double).
PowerShell only passes --index so Chinese never appears on the command line.
"""
from __future__ import annotations

import argparse
import json
import sys
import time
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))


def _save_retry(wb, path: Path, attempts: int = 10) -> None:
    last: Exception | None = None
    for i in range(attempts):
        try:
            wb.save(path)
            return
        except PermissionError as exc:
            last = exc
            time.sleep(1.2 + i * 0.4)
        except OSError as exc:
            last = exc
            time.sleep(1.2 + i * 0.4)
    raise RuntimeError(f"xlsx still locked after retries: {path}") from last


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", required=True, help="full export JSON from write_kanban_export_cfg")
    ap.add_argument("--index", required=True, type=int, help="city index in cfg.cities")
    ap.add_argument("--skip-register", action="store_true")
    args = ap.parse_args()

    cfg_path = Path(args.config)
    cfg = json.loads(cfg_path.read_text(encoding="utf-8"))
    cities = list(cfg.get("cities") or [])
    if args.index < 0 or args.index >= len(cities):
        print(f"city index out of range: {args.index} count={len(cities)}", file=sys.stderr)
        return 2

    city = str(cities[args.index])
    xlsx = Path(cfg["xlsx"])
    sheet = str(cfg.get("sheet") or "看板-单城")
    month = int(cfg["month"])
    region = str(cfg.get("region") or "")

    print(f"openpyxl filters index={args.index} month={month} city_len={len(city)}", flush=True)
    wb = load_workbook(xlsx, keep_links=False)
    try:
        if sheet not in wb.sheetnames:
            print(f"sheet missing: {sheet}; sheets={wb.sheetnames}", file=sys.stderr)
            return 2
        ws = wb[sheet]
        ws["C2"] = month
        ws["C3"] = city
        ws["E3"] = region
        _save_retry(wb, xlsx)
    finally:
        wb.close()

    city_cfg = dict(cfg)
    city_cfg["cities"] = [city]
    city_cfg["expectedCount"] = 1
    city_cfg["skipCellWrites"] = True
    city_cfg["skipRegister"] = bool(args.skip_register)
    out_dir = Path(cfg["outDir"])
    out_dir.mkdir(parents=True, exist_ok=True)
    city_cfg_path = out_dir / "_export_kanban_city.json"
    city_cfg_path.write_text(json.dumps(city_cfg, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"ok city_cfg={city_cfg_path}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
