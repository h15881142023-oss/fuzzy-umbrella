#!/usr/bin/env python3
"""Print which dates exist in 数据源(日) of an LR workbook."""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook  # noqa: E402

from config import CITIES, REGION_NAME  # noqa: E402
from lr.fill_template import DATA_SHEET, HEADER_ROW, _header_col_map, _read_row_day  # noqa: E402
from lr.table_utils import norm_header  # noqa: E402


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    args = ap.parse_args()
    path = Path(args.xlsx)
    if not path.exists():
        print(f"missing: {path}", file=sys.stderr)
        return 1

    wb = load_workbook(path, data_only=False)
    ws = wb[DATA_SHEET]
    col_map = _header_col_map(ws)
    day_col = col_map.get("日", 3)
    days: dict[str, set[str]] = {}
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        region = norm_header(ws.cell(r, col_map.get("区域", 1)).value)
        city = norm_header(ws.cell(r, col_map.get("组织结构", 2)).value)
        day = _read_row_day(ws.cell(r, day_col).value)
        if region == REGION_NAME and city in CITIES and day:
            days.setdefault(day.isoformat(), set()).add(city)
    wb.close()

    print(f"file={path}")
    print(f"days={len(days)}")
    for d in sorted(days):
        cities = ",".join(sorted(days[d]))
        print(f"  {d}: {len(days[d])} cities ({cities})")
    missing_full = [d for d, cs in days.items() if len(cs) < len(CITIES)]
    if missing_full:
        print(f"incomplete_days={missing_full}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
