"""将抓取结果写入 LR 模板「数据源(日)」（只写录入列，保留公式列）。"""
from __future__ import annotations

import shutil
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

from config import CITIES, REGION_NAME
from lr.table_utils import norm_header
from lr.xlsx_sanitize import sanitize_for_openpyxl

DATA_SHEET = "数据源(日)"
KANBAN_SHEET = "看板-单城"
TUANGOU_SHEET = "团购利润"
HEADER_ROW = 3
# 模板里 A–V（1–22）为网页录入列；其后多为公式/手工费用，不可覆盖
MAX_INPUT_COL = 22
FORMULA_TEMPLATE_ROW = 4
# 团购利润：仅仁寿/南溪；U 列每日利润固定值（覆盖模板公式）
TUANGOU_DAILY_PROFIT = {
    "仁寿县": -1500,
    "南溪": -1200,
}
TUANGOU_HEADER_ROW = 1
TUANGOU_COL_REGION = 1
TUANGOU_COL_CITY = 2
TUANGOU_COL_MONTH = 3
TUANGOU_COL_DAY = 4
TUANGOU_COL_PROFIT = 21  # U


def monthly_master_path(out_dir: Path, target: date) -> Path:
    return out_dir / f"LR日报_{target.year}-{target.month:02d}.xlsx"


def _latest_prior_workbook(out_dir: Path, target: date) -> Path | None:
    """同月、日期早于 target 的最近一份日报（用于累积填表）。"""
    best: Path | None = None
    best_day: date | None = None
    for path in out_dir.glob("LR日报_*.xlsx"):
        suffix = path.stem.removeprefix("LR日报_")
        if len(suffix) == 7 and suffix[4] == "-":
            continue
        try:
            day = date.fromisoformat(suffix)
        except ValueError:
            continue
        if day.year != target.year or day.month != target.month or day >= target:
            continue
        if best_day is None or day > best_day:
            best_day = day
            best = path
    return best


def _resolve_base_workbook(template: Path, out_dir: Path, target: date) -> Path:
    monthly = monthly_master_path(out_dir, target)
    if monthly.exists():
        return monthly
    prior = _latest_prior_workbook(out_dir, target)
    if prior and prior.exists():
        return prior
    return template


def count_filled_days(out_path: Path) -> int:
    """统计数据源(日)里已有多少个不同日期（五城齐算一天）。

    注意：read_only 模式禁止 ws.cell(r,c) 随机访问（会极慢/像卡死），必须 iter_rows。
    """
    wb = load_workbook(out_path, read_only=True, data_only=True)
    try:
        if DATA_SHEET not in wb.sheetnames:
            return 0
        ws = wb[DATA_SHEET]
        days: set[str] = set()
        region_i, city_i, day_i = 0, 1, 2
        for i, row in enumerate(
            ws.iter_rows(min_row=HEADER_ROW, max_col=MAX_INPUT_COL, values_only=True),
            start=HEADER_ROW,
        ):
            if i == HEADER_ROW:
                headers = [norm_header(c) for c in row]
                def _idx(name: str, default: int) -> int:
                    try:
                        return headers.index(name)
                    except ValueError:
                        return default
                region_i = _idx("区域", 0)
                city_i = _idx("组织结构", 1)
                day_i = _idx("日", 2)
                continue
            vals = list(row)

            def _get(idx: int):
                return vals[idx] if 0 <= idx < len(vals) else None

            region = norm_header(_get(region_i))
            city = norm_header(_get(city_i))
            day = _read_row_day(_get(day_i))
            if region == REGION_NAME and city in CITIES and day:
                days.add(day.isoformat())
        return len(days)
    finally:
        wb.close()


def _header_col_map(ws) -> dict[str, int]:
    """表头 → 列号；同名表头取首次出现（避免「日」映射到辅助列 AU）。"""
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(HEADER_ROW, col).value
        if not val:
            continue
        key = norm_header(val)
        if key not in mapping:
            mapping[key] = col
    return mapping


def _row_key(region: str, city: str, day: date) -> tuple[str, str, str]:
    return region, city, day.isoformat()


def _read_row_day(cell_val: Any) -> date | None:
    if isinstance(cell_val, datetime):
        return cell_val.date()
    if isinstance(cell_val, date):
        return cell_val
    if isinstance(cell_val, str):
        s = cell_val.strip()
        if s.startswith("="):
            return None
        if len(s) >= 10:
            try:
                return datetime.strptime(s[:10], "%Y-%m-%d").date()
            except ValueError:
                return None
    return None


def _copy_row_formulas(ws, src_row: int, dst_row: int) -> None:
    """把模板行的公式列复制到新行（含 AU/AV/AW 辅助列）。"""
    for col in range(MAX_INPUT_COL + 1, ws.max_column + 1):
        val = ws.cell(src_row, col).value
        if not (isinstance(val, str) and val.startswith("=")):
            continue
        origin = f"{get_column_letter(col)}{src_row}"
        dest = f"{get_column_letter(col)}{dst_row}"
        ws.cell(dst_row, col).value = Translator(val, origin=origin).translate_formula(dest)


def _set_kanban_month_city(wb, target: date, city: str = CITIES[0]) -> None:
    if KANBAN_SHEET not in wb.sheetnames:
        return
    ws = wb[KANBAN_SHEET]
    ws["C2"] = target.month
    ws["C3"] = city
    ws["E3"] = REGION_NAME


def _tuangou_day_label(val: Any) -> str | None:
    """统一为「M月D日」文本，便于与新表展示一致。"""
    if isinstance(val, datetime):
        return f"{val.month}月{val.day}日"
    if isinstance(val, date):
        return f"{val.month}月{val.day}日"
    if isinstance(val, str):
        s = val.replace("\n", "").replace(" ", "").strip()
        return s or None
    return None


def _last_used_row(ws, col: int = 1) -> int:
    last = 1
    for r in range(1, (ws.max_row or 1) + 1):
        if ws.cell(r, col).value is not None:
            last = r
    return last


def _fill_tuangou_profit(wb, target: date) -> None:
    """写入团购利润：A 区域、B 城市、C 月份、D 日期、U 每日利润（仁寿-1500/南溪-1200）。"""
    if TUANGOU_SHEET not in wb.sheetnames:
        print(f"[lr] skip {TUANGOU_SHEET}: sheet missing", flush=True)
        return
    ws = wb[TUANGOU_SHEET]
    month_num = int(f"{target.year}{target.month:02d}")
    day_label = f"{target.month}月{target.day}日"

    index: dict[tuple[str, str], int] = {}
    for r in range(TUANGOU_HEADER_ROW + 1, (ws.max_row or 1) + 1):
        city = norm_header(ws.cell(r, TUANGOU_COL_CITY).value)
        label = _tuangou_day_label(ws.cell(r, TUANGOU_COL_DAY).value)
        if city and label:
            index[(city, label)] = r

    next_row = _last_used_row(ws, TUANGOU_COL_REGION) + 1
    for city, profit in TUANGOU_DAILY_PROFIT.items():
        key = (city, day_label)
        row_idx = index.get(key)
        if row_idx is None:
            row_idx = next_row
            next_row += 1
        ws.cell(row_idx, TUANGOU_COL_REGION, REGION_NAME)
        ws.cell(row_idx, TUANGOU_COL_CITY, city)
        # 月份必须是 202608 这种数字，不能被单元格日期格式吞成序列日
        mcell = ws.cell(row_idx, TUANGOU_COL_MONTH, month_num)
        mcell.number_format = "0"
        dcell = ws.cell(row_idx, TUANGOU_COL_DAY, day_label)
        dcell.number_format = "@"
        pcell = ws.cell(row_idx, TUANGOU_COL_PROFIT, profit)
        pcell.number_format = "General"
        print(
            f"[lr] 团购利润 row={row_idx} {city} {day_label} month={month_num} U={profit}",
            flush=True,
        )


def _sanitized_template_cache(template: Path, out_dir: Path) -> Path:
    """首次清洗模板较慢；缓存到 work，后续天直接复制。"""
    cache = out_dir / "_template_sanitized.xlsx"
    try:
        if cache.exists() and cache.stat().st_mtime >= template.stat().st_mtime:
            return cache
    except OSError:
        pass
    print("[lr] building sanitized template cache (once, may take 1-3 min) ...", flush=True)
    sanitize_for_openpyxl(template, cache)
    return cache


def fill_template(
    template: Path,
    rows: list[dict[str, Any]],
    target: date,
    out_dir: Path,
) -> Path:
    """写入目标日五城；同月累积：在已有 workbook 上追加/更新，而非每天从空白模板复制。"""
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"LR日报_{target.isoformat()}.xlsx"
    base = _resolve_base_workbook(template, out_dir, target)
    print(f"[lr] fill base={base.name} -> {out_path.name}", flush=True)
    if base.resolve() == template.resolve():
        cached = _sanitized_template_cache(template, out_dir)
        print(f"[lr] copy sanitized cache -> {out_path.name}", flush=True)
        shutil.copy2(cached, out_path)
    else:
        print(f"[lr] copy prior workbook {base.name}", flush=True)
        shutil.copy2(base, out_path)

    print("[lr] openpyxl load_workbook (large xlsx, may take 1-5 min) ...", flush=True)
    wb = load_workbook(out_path, keep_links=False)
    print("[lr] workbook loaded, writing rows ...", flush=True)
    if DATA_SHEET not in wb.sheetnames:
        raise ValueError(f"模板缺少工作表: {DATA_SHEET}")
    ws = wb[DATA_SHEET]
    col_map = _header_col_map(ws)

    day_col = col_map.get("日", 3)
    if day_col != 3:
        # 防御：若仍映射错误，强制用 C 列
        day_col = 3
        col_map["日"] = 3

    index: dict[tuple[str, str, str], int] = {}
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        region = norm_header(ws.cell(r, col_map.get("区域", 1)).value)
        city = norm_header(ws.cell(r, col_map.get("组织结构", 2)).value)
        day = _read_row_day(ws.cell(r, day_col).value)
        if region and city and day:
            index[_row_key(region, city, day)] = r

    by_city = {str(r.get("组织结构")): r for r in rows if r.get("组织结构") in CITIES}
    missing = [c for c in CITIES if c not in by_city]
    if missing:
        raise ValueError(f"抓取数据缺少城市: {missing}；目标日={target}")

    next_row = ws.max_row + 1
    for city in CITIES:
        src = by_city[city]
        key = _row_key(REGION_NAME, city, target)
        is_new = key not in index
        row_idx = index.get(key, next_row)
        if is_new:
            _copy_row_formulas(ws, FORMULA_TEMPLATE_ROW, row_idx)
            next_row += 1

        for header, val in src.items():
            col = col_map.get(norm_header(header))
            if not col or col > MAX_INPUT_COL:
                continue
            if header == "日" or norm_header(header) == "日":
                val = datetime.combine(target, datetime.min.time())
            ws.cell(row_idx, col, val)

        # 保证关键三列正确
        ws.cell(row_idx, col_map.get("区域", 1), REGION_NAME)
        ws.cell(row_idx, col_map.get("组织结构", 2), city)
        ws.cell(row_idx, day_col, datetime.combine(target, datetime.min.time()))

    _fill_tuangou_profit(wb, target)
    _set_kanban_month_city(wb, target)
    print("[lr] saving workbook ...", flush=True)
    wb.save(out_path)
    wb.close()

    monthly = monthly_master_path(out_dir, target)
    shutil.copy2(out_path, monthly)
    print(f"[lr] fill done -> {out_path.name} (+ {monthly.name})", flush=True)
    return out_path
