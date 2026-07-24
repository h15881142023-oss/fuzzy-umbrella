"""将「看板-单城」导出为五城 PNG（Windows 优先 Python COM + 剪贴板）。"""
from __future__ import annotations

import os
import subprocess
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from config import CITIES, REGION_NAME

KANBAN_SHEET = "看板-单城"
EXPORT_PS1 = Path(__file__).resolve().parent.parent / "scripts" / "export_lr_kanban_wps.ps1"


def _render_with_pillow(xlsx: Path, out_png: Path, city: str | None = None) -> Path:
    from PIL import Image, ImageDraw, ImageFont

    wb = load_workbook(xlsx, data_only=True)
    ws = wb[KANBAN_SHEET]
    lines: list[str] = []
    title = ws.cell(1, 2).value or "看板-单城"
    if city:
        lines.append(f"{city} | {title}")
    else:
        lines.append(str(title))
    lines.append("")

    for r in range(2, min(ws.max_row, 40) + 1):
        cells = []
        for c in range(2, min(ws.max_column, 18) + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            if hasattr(v, "strftime"):
                v = v.strftime("%Y-%m-%d")
            elif isinstance(v, float):
                v = f"{v:.4g}" if abs(v) < 1 else f"{v:.2f}"
            cells.append(str(v))
        if cells:
            lines.append(" | ".join(cells))
    wb.close()

    font = ImageFont.load_default()
    line_h = 18
    width = 1400
    height = max(400, 40 + line_h * len(lines))
    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)
    y = 16
    for i, line in enumerate(lines):
        draw.text((16, y), line[:220], fill=(0, 80, 160) if i == 0 else (20, 20, 20), font=font)
        y += line_h

    out_png.parent.mkdir(parents=True, exist_ok=True)
    img.save(out_png, format="PNG")
    return out_png


def export_kanban_pngs_wps_ps1(
    xlsx: Path,
    out_dir: Path,
    target: date,
    cities: list[str] | None = None,
) -> list[Path]:
    """Fallback：PowerShell COM 脚本（参数经 UTF-8 JSON，避免命令行乱码）。"""
    import json

    cities = cities or list(CITIES)
    if not EXPORT_PS1.exists():
        raise FileNotFoundError(f"missing export script: {EXPORT_PS1}")
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_abs = xlsx.resolve()
    out_abs = out_dir.resolve()
    log_path = out_abs / "wps_export.log"
    cfg_path = out_abs / "_export_kanban_cfg.json"
    cfg = {
        "xlsx": str(xlsx_abs),
        "outDir": str(out_abs),
        "month": target.month,
        "region": REGION_NAME,
        "cities": cities,
        "sheet": KANBAN_SHEET,
        "range": "B1:R37",
    }
    cfg_path.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")

    # ASCII-only wrapper to avoid PS parse errors on GBK systems
    wrapper = out_abs / "_export_kanban_once.ps1"
    wrapper.write_text(
        f'& "{EXPORT_PS1}" -ConfigJson "{cfg_path}"\n',
        encoding="ascii",
        errors="strict",
    )
    cmdline = (
        f'chcp 65001 >NUL & powershell.exe -NoProfile -ExecutionPolicy Bypass '
        f'-File "{wrapper}" > "{log_path}" 2>&1'
    )
    proc = subprocess.run(
        ["cmd.exe", "/c", cmdline],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=600,
    )
    log_text = ""
    if log_path.exists():
        log_text = log_path.read_text(encoding="utf-8", errors="replace")
        if not log_text.strip():
            log_text = log_path.read_text(encoding="gbk", errors="replace")
    if proc.returncode != 0:
        raise RuntimeError(
            f"WPS PS1 export failed code={proc.returncode}\nLOG:\n{log_text[-4000:]}"
        )
    pngs: list[Path] = []
    for city in cities:
        safe = "".join("_" if c in '\\/:*?"<>|' else c for c in city)
        cand = out_dir / f"看板-单城_{safe}_{target.month}.png"
        if not cand.exists():
            matches = sorted(out_dir.glob(f"看板-单城_{safe}_*.png"))
            if not matches:
                raise FileNotFoundError(f"PNG missing for {city}\nLOG:\n{log_text[-4000:]}")
            cand = matches[-1]
        pngs.append(cand)
    return pngs


def export_kanban_pngs_wps(
    xlsx: Path,
    out_dir: Path,
    target: date,
    cities: list[str] | None = None,
) -> list[Path]:
    """Windows：优先 Python COM + 剪贴板；失败再回退 PowerShell。"""
    cities = cities or list(CITIES)
    errors: list[str] = []

    try:
        from lr.export_kanban_com import export_kanban_pngs_com

        print("[lr] try Python COM + clipboard export ...", flush=True)
        return export_kanban_pngs_com(xlsx, out_dir, target, cities)
    except Exception as exc:  # noqa: BLE001
        errors.append(f"com: {exc}")
        print(f"[lr] COM export failed: {exc}", flush=True)

    try:
        print("[lr] fallback PowerShell export ...", flush=True)
        return export_kanban_pngs_wps_ps1(xlsx, out_dir, target, cities)
    except Exception as exc:  # noqa: BLE001
        errors.append(f"ps1: {exc}")

    raise RuntimeError("Kanban PNG export failed. " + " || ".join(errors))


def export_kanban_pngs(
    xlsx: Path,
    out_dir: Path,
    target: date,
    cities: list[str] | None = None,
    *,
    allow_pillow_fallback: bool = False,
) -> list[Path]:
    """导出五城看板图。Windows 走 WPS/Excel；其它平台仅在允许时用 Pillow 占位。"""
    cities = cities or list(CITIES)
    out_dir.mkdir(parents=True, exist_ok=True)

    if sys.platform == "win32" or os.environ.get("LR_FORCE_WPS") == "1":
        return export_kanban_pngs_wps(xlsx, out_dir, target, cities)

    if not allow_pillow_fallback:
        raise RuntimeError(
            "当前环境无法调用 WPS/Excel 截图。请在 Windows 本机运行，"
            "或设置 allow_pillow_fallback / 环境变量 LR_ALLOW_PILLOW=1（仅调试）。"
        )

    pngs: list[Path] = []
    for city in cities:
        wb = load_workbook(xlsx)
        ws = wb[KANBAN_SHEET]
        ws["C2"] = target.month
        ws["C3"] = city
        ws["E3"] = REGION_NAME
        wb.save(xlsx)
        wb.close()
        png = out_dir / f"看板-单城_{city}_{target.isoformat()}.png"
        _render_with_pillow(xlsx, png, city=city)
        pngs.append(png)
    return pngs


def _render_with_libreoffice(xlsx: Path, out_dir: Path) -> Path | None:
    for cmd in ("soffice", "libreoffice"):
        try:
            subprocess.run(
                [cmd, "--headless", "--convert-to", "png", "--outdir", str(out_dir), str(xlsx)],
                check=True,
                capture_output=True,
                timeout=120,
            )
            pngs = sorted(out_dir.glob("*.png"), key=lambda p: p.stat().st_mtime, reverse=True)
            return pngs[0] if pngs else None
        except (FileNotFoundError, subprocess.CalledProcessError, subprocess.TimeoutExpired):
            continue
    return None


def export_kanban_png(xlsx: Path, out_png: Path) -> Path:
    """兼容旧接口：导出单张（Pillow/LibreOffice，不切换城市）。"""
    png = _render_with_libreoffice(xlsx, out_png.parent)
    if png and png.exists():
        if png != out_png:
            png.replace(out_png)
        return out_png
    return _render_with_pillow(xlsx, out_png)
