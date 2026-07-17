"""解析「外卖陪访拜访看板」后台导出的拜访 Excel → visit_check payload。"""
from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, BinaryIO

from openpyxl import load_workbook

from config import CITIES, normalize_city
from scrapers.visit_check import SHEET_TO_CITY

# 标准城市名 → 检核引擎 sheet 键（与历史金山文档 sheet 名对齐）
CITY_TO_SHEET = {v: k for k, v in SHEET_TO_CITY.items()}

COL_ALIASES = {
    "city": ("城市",),
    "target": ("拜访对象",),
    "type": ("拜访对象类型",),
    "time": ("拜访时间",),
    "bd": ("拜访BD名字", "拜访BD姓名", "BD名字"),
    "desc": ("工作描述",),
    "region": ("区域",),
}


def _header_map(headers: list[Any]) -> dict[str, int]:
    idx: dict[str, int] = {}
    for i, h in enumerate(headers):
        name = str(h or "").strip()
        for key, aliases in COL_ALIASES.items():
            if name in aliases:
                idx[key] = i
                break
    required = ("city", "target", "type", "time", "bd", "desc")
    missing = [k for k in required if k not in idx]
    if missing:
        raise ValueError(f"Excel 缺少列: {missing}；实际表头={headers}")
    return idx


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    return str(v).strip()


def _infer_check_date(times: list[str]) -> str:
    dates = []
    for t in times:
        s = (t or "")[:10].replace("/", "-")
        if len(s) >= 10:
            dates.append(s[:10])
    if dates:
        return max(dates)
    return (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")


def excel_to_payload(source: str | Path | BinaryIO) -> dict[str, Any]:
    """读后台导出 xlsx，返回 check_payload 所需结构。"""
    if hasattr(source, "read"):
        wb = load_workbook(source, read_only=True, data_only=True)
    else:
        wb = load_workbook(Path(source), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        raise ValueError("Excel 为空")

    col = _header_map(list(rows[0]))
    by_city: dict[str, list[dict]] = defaultdict(list)
    all_times: list[str] = []

    for row in rows[1:]:
        if not row:
            continue
        city_raw = _cell_str(row[col["city"]])
        city = normalize_city(city_raw)
        if not city or city not in CITIES:
            continue
        time_s = _cell_str(row[col["time"]])
        if not time_s:
            continue
        all_times.append(time_s)
        by_city[city].append(
            {
                "target": _cell_str(row[col["target"]]),
                "type": _cell_str(row[col["type"]]),
                "time": time_s,
                "bd": _cell_str(row[col["bd"]]),
                "desc": _cell_str(row[col["desc"]]),
            }
        )

    check_date = _infer_check_date(all_times)
    cities_out: dict[str, dict] = {}
    for city in CITIES:
        sheet = CITY_TO_SHEET.get(city, city)
        records = by_city.get(city) or []
        cities_out[sheet] = {
            "records": records,
            "t1Count": len(records),
            "totalRows": len(records),
        }

    return {
        "source": "admin_export",
        "targetDate": check_date,
        "check_date": check_date,
        "cities": cities_out,
    }
