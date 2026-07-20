import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from config import CITIES
from lr.fill_template import DATA_SHEET, HEADER_ROW, _header_col_map, fill_template


class HeaderColumnMapTest(unittest.TestCase):
    def test_uses_first_duplicate_header(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(HEADER_ROW, 3, "日")
        sheet.cell(HEADER_ROW, 47, "日")

        mapping = _header_col_map(sheet)

        self.assertEqual(mapping["日"], 3)
        workbook.close()

    def test_fills_preformatted_rows_and_preserves_formulas(self) -> None:
        target = date(2026, 7, 19)
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            template = root / "template.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = DATA_SHEET
            for col, header in enumerate(
                ["区域", "组织结构", "日", "原价交易额"],
                start=1,
            ):
                sheet.cell(HEADER_ROW, col, header)
            for row in range(4, 9):
                sheet.cell(row, 1, f"=VLOOKUP(B{row},基础档案!C:E,3,0)")
                sheet.cell(row, 23, f"=D{row}*2")
            workbook.save(template)
            workbook.close()

            rows = [
                {
                    "区域": "川藏一区",
                    "组织结构": city,
                    "日": target,
                    "原价交易额": index,
                }
                for index, city in enumerate(CITIES, start=1)
            ]
            output = fill_template(template, rows, target, root / "output")

            result = load_workbook(output, data_only=False)
            result_sheet = result[DATA_SHEET]
            for index, city in enumerate(CITIES, start=1):
                row = index + 3
                self.assertEqual(result_sheet.cell(row, 1).value, "川藏一区")
                self.assertEqual(result_sheet.cell(row, 2).value, city)
                self.assertEqual(
                    result_sheet.cell(row, 3).value,
                    datetime(2026, 7, 19),
                )
                self.assertEqual(result_sheet.cell(row, 4).value, index)
                self.assertEqual(result_sheet.cell(row, 23).value, f"=D{row}*2")
            result.close()


if __name__ == "__main__":
    unittest.main()
